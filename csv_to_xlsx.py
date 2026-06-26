"""Convert a submission CSV to a formatted XLSX workbook.

The official Redrob submission must remain CSV (the validator rejects .xlsx);
this produces an XLSX *copy* for presentation / alternate-format needs.

Usage:
    python csv_to_xlsx.py --in submission.csv --out submission.xlsx
"""

from __future__ import annotations

import argparse
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def convert(in_path: str, out_path: str, sheet_name: str = "Ranked Candidates") -> int:
    """Read a submission CSV and write a styled XLSX. Returns the row count."""
    with open(in_path, "r", encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c)
            # Type the rank as int and score as float for real spreadsheet use.
            if r > 1 and c == 2:
                cell.value = int(value)
            elif r > 1 and c == 3:
                cell.value = float(value)
            else:
                cell.value = value
            if r == 1:
                cell.fill = header_fill
                cell.font = header_font

    # Column widths: id / rank / score compact, reasoning wide.
    widths = {1: 16, 2: 8, 3: 12, 4: 90}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Wrap the reasoning column.
    for r in range(2, len(rows) + 1):
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    ws.auto_filter.ref = f"A1:D{len(rows)}"

    wb.save(out_path)
    return max(0, len(rows) - 1)


def main() -> None:
    p = argparse.ArgumentParser(description="Convert submission CSV to XLSX.")
    p.add_argument("--in", dest="in_path", default="submission.csv")
    p.add_argument("--out", dest="out_path", default="submission.xlsx")
    args = p.parse_args()
    n = convert(args.in_path, args.out_path)
    print(f"Wrote {args.out_path} with {n} ranked rows (+ header).")


if __name__ == "__main__":
    main()
